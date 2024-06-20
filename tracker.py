import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import pytesseract
from PIL import Image
import re
import os

# File name for storing data
EXCEL_FILE = 'gym_progress.xlsx'

# Initialize Excel file
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Exercise', 'Weight', 'Sets', 'Reps', 'Distance', 'Duration'])
    wb.save(EXCEL_FILE)

# Save data to Excel file
def save_data(date, exercise, weight, sets, reps, distance, duration):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([date, exercise, weight, sets, reps, distance, duration])
    wb.save(EXCEL_FILE)
    messagebox.showinfo("Data Saved", "Your workout data has been saved.")

# Calculate statistics
def calculate_statistics():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    total_distance = 0
    total_duration = 0
    run_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, exercise, weight, sets, reps, distance, duration = row
        if exercise in ['Running', 'Walking', 'Cycling']:
            total_distance += distance if distance else 0
            if duration:
                h, m, s = map(int, duration.split(':'))
                total_duration += h * 3600 + m * 60 + s
            run_count += 1
    
    if run_count > 0:
        average_pace = total_duration / total_distance if total_distance > 0 else 0
        stats_label.config(text=f"You have run a total of {total_distance:.2f} miles so far. Keep it up!\n"
                                f"Average pace: {average_pace/60:.2f} min/mile\n"
                                f"Total time: {total_duration//3600}h {(total_duration%3600)//60}m {total_duration%60}s")
    else:
        stats_label.config(text="No running data available.")

# Extract data from handwritten notes using OCR
def process_ocr_image(image_path):
    text = pytesseract.image_to_string(Image.open(image_path))
    entries = re.findall(r'(\d{2}/\d{2}/\d{2}).*?(Running|Walking|Cycling|Gym).*?(\d+\.?\d*).*?(\d+:\d+:\d+)', text, re.DOTALL)
    
    for entry in entries:
        date, exercise, distance, duration = entry
        save_data(date, exercise, None, None, None, float(distance), duration)
    
    messagebox.showinfo("Data Processed", "Handwritten data has been processed and saved.")

# GUI
root = tk.Tk()
root.title("Gym Progress Tracker")

# Left frame for adding data
left_frame = tk.Frame(root, padx=20, pady=20, bg="#5b9aa0")
left_frame.grid(row=0, column=0, padx=10, pady=10, sticky="n")

tk.Label(left_frame, text="Add Data", font=("Arial", 16, "bold"), bg="#5b9aa0", fg="white").grid(row=0, columnspan=2)
tk.Label(left_frame, text="Enter info below:", font=("Arial", 12), bg="#5b9aa0", fg="white").grid(row=1, columnspan=2)

# Date
tk.Label(left_frame, text="Date", bg="#5b9aa0", fg="white").grid(row=2, column=0, sticky="e")
date_entry = tk.Entry(left_frame)
date_entry.grid(row=2, column=1)

# Distance
tk.Label(left_frame, text="Distance", bg="#5b9aa0", fg="white").grid(row=3, column=0, sticky="e")
distance_entry = tk.Entry(left_frame)
distance_entry.grid(row=3, column=1)

# Duration
tk.Label(left_frame, text="Duration", bg="#5b9aa0", fg="white").grid(row=4, column=0, sticky="e")
duration_entry = tk.Entry(left_frame)
duration_entry.grid(row=4, column=1)

# Submit Button
submit_button = tk.Button(left_frame, text="Submit", command=lambda: save_data(
    date_entry.get(), 'Running', None, None, None, float(distance_entry.get()), duration_entry.get()))
submit_button.grid(row=5, columnspan=2, pady=10)

# Right frame for statistics
right_frame = tk.Frame(root, padx=20, pady=20, bg="#5b9aa0")
right_frame.grid(row=0, column=1, padx=10, pady=10, sticky="n")

tk.Label(right_frame, text="See Running Statistics", font=("Arial", 16, "bold"), bg="#5b9aa0", fg="white").grid(row=0, columnspan=2)
tk.Label(right_frame, text="Analysis options:", font=("Arial", 12), bg="#5b9aa0", fg="white").grid(row=1, columnspan=2)

# Analysis Buttons
tk.Button(right_frame, text="Total Distance", command=calculate_statistics).grid(row=2, column=0, padx=5, pady=5)
tk.Button(right_frame, text="Average Pace", command=calculate_statistics).grid(row=2, column=1, padx=5, pady=5)
tk.Button(right_frame, text="Total Time", command=calculate_statistics).grid(row=3, columnspan=2, padx=5, pady=5)

# Statistics Label
stats_label = tk.Label(right_frame, text="Stats will be displayed here.", bg="#5b9aa0", fg="white", justify="left", font=("Arial", 12))
stats_label.grid(row=4, columnspan=2, pady=10)

# Upload and process handwritten notes
def upload_image():
    file_path = filedialog.askopenfilename()
    if file_path:
        process_ocr_image(file_path)

# Upload Image Button
upload_button = tk.Button(left_frame, text="Upload Handwritten Notes", command=upload_image)
upload_button.grid(row=6, columnspan=2, pady=10)

root.mainloop()
